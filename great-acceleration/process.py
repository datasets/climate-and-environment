"""
Process IGBP Great Acceleration Excel data into clean CSV files.
Source: IGBP Great Acceleration data collection (October 2014)
Paper: Steffen et al. 2015, The Anthropocene Review
       "The Trajectory of the Anthropocene: The Great Acceleration"
       DOI: 10.1177/2053019614564785
"""
import csv
import openpyxl

EXCEL_FILE = "great-acceleration-data.xlsx"

# ── Sheet metadata ────────────────────────────────────────────────────────────

SOCIOECONOMIC_SHEETS = [
    {
        "sheet": "1 Population",
        "id": "population",
        "name": "Population",
        "unit": "billion",
        "columns": ["oecd", "brics", "rest", "world"],
    },
    {
        "sheet": "2 Real GDP",
        "id": "real-gdp",
        "name": "Real GDP",
        "unit": "trillion USD (2005)",
        "columns": ["oecd", "brics", "rest", "world"],
    },
    {
        "sheet": "3 FDI",
        "id": "fdi",
        "name": "Foreign Direct Investment",
        "unit": "trillion USD (current)",
        "columns": ["oecd", "brics", "rest", "world"],
    },
    {
        "sheet": "4 Urban population",
        "id": "urban-population",
        "name": "Urban Population",
        "unit": "billion",
        "columns": ["oecd", "brics", "rest", "world"],
    },
    {
        "sheet": "5 Primary energy use",
        "id": "primary-energy",
        "name": "Primary Energy Use",
        "unit": "exajoule (EJ)",
        "columns": ["world"],
    },
    {
        "sheet": "6 Fertilizer consumption",
        "id": "fertilizer",
        "name": "Fertilizer Consumption",
        "unit": "million tonnes",
        "columns": ["oecd", "brics", "rest", "world"],
    },
    {
        "sheet": "7 Large dams",
        "id": "large-dams",
        "name": "Large Dams",
        "unit": "thousand dams (cumulative)",
        "columns": ["oecd", "brics", "rest", "world"],
    },
    {
        "sheet": "8 Water use",
        "id": "water-use",
        "name": "Water Use",
        "unit": "thousand km³",
        "columns": ["oecd", "brics", "rest", "world"],
    },
    {
        "sheet": "9 Paper production",
        "id": "paper-production",
        "name": "Paper Production",
        "unit": "million tonnes",
        "columns": ["oecd", "brics", "rest", "world"],
    },
    {
        "sheet": "10 Transportation",
        "id": "transportation",
        "name": "Transportation (Motor Vehicles)",
        "unit": "million vehicles",
        "columns": ["oecd", "brics", "rest", "world"],
    },
    {
        "sheet": "11 Telecommunications",
        "id": "telecommunications",
        "name": "Telecommunications",
        "unit": "billion subscriptions (landlines + mobile)",
        "columns": ["oecd", "brics", "rest", "world"],
    },
    {
        "sheet": "12 International Tourism",
        "id": "international-tourism",
        "name": "International Tourism",
        "unit": "million arrivals",
        "columns": ["world"],
    },
]

EARTH_SYSTEM_SHEETS = [
    {
        "sheet": "1 CarbonDioxide",
        "id": "co2",
        "name": "Carbon Dioxide",
        "unit": "ppm",
    },
    {
        "sheet": "2 NitrousOxide",
        "id": "n2o",
        "name": "Nitrous Oxide",
        "unit": "ppb",
    },
    {
        "sheet": "3 Methane",
        "id": "ch4",
        "name": "Methane",
        "unit": "ppb",
    },
    {
        "sheet": "4 Ozone",
        "id": "ozone",
        "name": "Stratospheric Ozone",
        "unit": "% loss",
    },
    {
        "sheet": "5 Temperature",
        "id": "temperature",
        "name": "Surface Temperature Anomaly",
        "unit": "°C (relative to 1961-1990)",
    },
    {
        "sheet": "6 OceanAcidification",
        "id": "ocean-acidification",
        "name": "Ocean Acidification",
        "unit": "nmol/kg (H+ concentration)",
    },
    {
        "sheet": "7 Marine fish",
        "id": "marine-fish",
        "name": "Marine Fish Capture",
        "unit": "million tonnes",
    },
    {
        "sheet": "8 ShrimpAqu",
        "id": "shrimp-aquaculture",
        "name": "Shrimp Aquaculture",
        "unit": "million tonnes",
    },
    {
        "sheet": "9 Nitrogen",
        "id": "nitrogen",
        "name": "Nitrogen to Coastal Zone",
        "unit": "million tonnes/year",
    },
    {
        "sheet": "10 TropicalForest",
        "id": "tropical-forest",
        "name": "Tropical Forest Loss",
        "unit": "%",
    },
    {
        "sheet": "11 DomLand",
        "id": "domesticated-land",
        "name": "Domesticated Land",
        "unit": "% of total land area",
    },
    {
        "sheet": "12 Terrestrial biosph degradati",
        "id": "terrestrial-biosphere",
        "name": "Terrestrial Biosphere Degradation",
        "unit": "% decrease in mean species abundance",
    },
]


# ── Helpers ───────────────────────────────────────────────────────────────────

def find_data_start(rows):
    """Find the index of the first row with numeric year + at least one numeric value."""
    for i, row in enumerate(rows):
        if row and isinstance(row[0], (int, float)):
            # Check if any other column has a numeric value
            rest = [x for x in row[1:] if isinstance(x, (int, float))]
            if rest:
                return i
    return None


def find_header_row(rows):
    """Find header row index (contains 'year' or 'year ad' in first cell)."""
    for i, row in enumerate(rows):
        non_none = [x for x in row if x is not None]
        if non_none and isinstance(non_none[0], str):
            if non_none[0].strip().lower() in ("year", "year ad"):
                return i
    return None


def safe_float(val):
    """Convert value to float or return None."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    if isinstance(val, str) and val.strip().upper() in ("#N/A", "N/A", "", "NA"):
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


# ── Process socioeconomic data ────────────────────────────────────────────────

def process_socioeconomic(wb):
    """Extract socioeconomic indicators into long-format rows."""
    rows_out = []

    for meta in SOCIOECONOMIC_SHEETS:
        ws = wb[meta["sheet"]]
        rows = list(ws.iter_rows(values_only=True))

        # Find header to determine column order
        header_idx = find_header_row(rows)
        if header_idx is None:
            print(f"  WARNING: No header found in {meta['sheet']}")
            continue

        header = [str(x).strip().lower() if x else "" for x in rows[header_idx]]

        # Map column positions
        col_map = {}
        for j, h in enumerate(header):
            if "year" in h:
                col_map["year"] = j
            elif "oecd" in h or "oedc" in h:  # handle typo "oedc" in fertilizer sheet
                col_map["oecd"] = j
            elif "brics" in h:
                col_map["brics"] = j
            elif "rest" in h:
                col_map["rest"] = j
            elif "world" in h:
                col_map["world"] = j

        # For single-value sheets (no world/oecd/brics), treat second column as world
        if "world" not in col_map and "oecd" not in col_map:
            year_col = col_map.get("year", 0)
            # Find the first non-year column with data
            for j, h in enumerate(header):
                if j != year_col and h:
                    col_map["world"] = j
                    break

        # Find data rows (must start after header)
        data_start = header_idx + 1
        for i in range(data_start, len(rows)):
            row = rows[i]
            if not row:
                continue

            year_val = row[col_map.get("year", 0)] if "year" in col_map else None
            if not isinstance(year_val, (int, float)):
                continue
            year = int(year_val)

            world = safe_float(row[col_map["world"]]) if "world" in col_map else None
            oecd = safe_float(row[col_map["oecd"]]) if "oecd" in col_map else None
            brics = safe_float(row[col_map["brics"]]) if "brics" in col_map else None
            rest = safe_float(row[col_map["rest"]]) if "rest" in col_map else None

            # Skip rows with no useful data
            if all(v is None for v in [world, oecd, brics, rest]):
                continue

            rows_out.append({
                "year": year,
                "indicator_id": meta["id"],
                "indicator": meta["name"],
                "unit": meta["unit"],
                "world": world,
                "oecd": oecd,
                "brics": brics,
                "rest": rest,
            })

    return rows_out


# ── Process earth system data ─────────────────────────────────────────────────

def process_earth_system(wb):
    """Extract earth system indicators into long-format rows."""
    rows_out = []

    for meta in EARTH_SYSTEM_SHEETS:
        ws = wb[meta["sheet"]]
        rows = list(ws.iter_rows(values_only=True))

        # Find header row
        header_idx = find_header_row(rows)
        if header_idx is None:
            print(f"  WARNING: No header found in {meta['sheet']}")
            continue

        data_start = header_idx + 1
        for i in range(data_start, len(rows)):
            row = rows[i]
            if not row:
                continue

            year_val = row[0]
            if not isinstance(year_val, (int, float)):
                continue
            year = int(year_val)

            # Value is in second column
            value = safe_float(row[1]) if len(row) > 1 else None
            if value is None:
                continue

            rows_out.append({
                "year": year,
                "indicator_id": meta["id"],
                "indicator": meta["name"],
                "unit": meta["unit"],
                "value": value,
            })

    return rows_out


# ── Write CSVs ────────────────────────────────────────────────────────────────

def write_socioeconomic(rows):
    path = "data/socioeconomic.csv"
    fieldnames = ["year", "indicator_id", "indicator", "unit", "world", "oecd", "brics", "rest"]
    with open(path, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for row in sorted(rows, key=lambda r: (r["indicator_id"], r["year"])):
            writer.writerow(row)
    print(f"  Wrote {len(rows):,} rows to {path}")


def write_earth_system(rows):
    path = "data/earth-system.csv"
    fieldnames = ["year", "indicator_id", "indicator", "unit", "value"]
    with open(path, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for row in sorted(rows, key=lambda r: (r["indicator_id"], r["year"])):
            writer.writerow(row)
    print(f"  Wrote {len(rows):,} rows to {path}")


# ── Main ──────────────────────────────────────────────────────────────────────

import warnings
warnings.filterwarnings("ignore")

wb = openpyxl.load_workbook(EXCEL_FILE, read_only=True)

print("Processing socioeconomic indicators...")
socio_rows = process_socioeconomic(wb)
write_socioeconomic(socio_rows)

print("Processing earth system indicators...")
earth_rows = process_earth_system(wb)
write_earth_system(earth_rows)

# Print summary
print("\nSummary:")
print(f"  Socioeconomic: {len(socio_rows):,} rows across {len(SOCIOECONOMIC_SHEETS)} indicators")
print(f"  Earth system:  {len(earth_rows):,} rows across {len(EARTH_SYSTEM_SHEETS)} indicators")

# Show year ranges per indicator
all_rows = [(r["indicator_id"], r["year"]) for r in socio_rows + earth_rows]
from collections import defaultdict
by_indicator = defaultdict(list)
for ind_id, year in all_rows:
    by_indicator[ind_id].append(year)

print("\nYear ranges per indicator:")
for ind_id in sorted(by_indicator):
    years = by_indicator[ind_id]
    print(f"  {ind_id}: {min(years)}-{max(years)} ({len(years)} data points)")
