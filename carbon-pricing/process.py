"""
process.py — Carbon Pricing & Emissions Trading Schemes

Sources:
  1. World Bank Carbon Pricing Dashboard (compliance initiatives, prices, revenue)
     https://carbonpricingdashboard.worldbank.org/
  2. New Zealand ETS monthly NZU prices
     https://github.com/theecanmole/nzu

Outputs:
  data/carbon-pricing-initiatives.csv  — overview of all compliance CPIs
  data/carbon-prices.csv               — annual carbon prices (long format)
  data/nz-ets-monthly-prices.csv       — NZ ETS monthly NZU prices 2010–present
"""

import csv
import io
import urllib.request

import openpyxl

WB_URL = "https://carbonpricingdashboard.worldbank.org/sites/default/files/data-latest.xlsx"
NZU_URL = "https://raw.githubusercontent.com/theecanmole/nzu/master/nzu-month-price.csv"


def download_bytes(url):
    with urllib.request.urlopen(url) as r:
        return r.read()


def parse_wb_excel(data):
    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True)

    # --- Sheet 1: Compliance_Gen Info → initiatives overview ---
    ws_gen = wb["Compliance_Gen Info"]
    gen_rows = list(ws_gen.iter_rows(values_only=True))
    # Row 0: metadata note; Row 1: headers; Row 2+: data
    gen_headers = gen_rows[1]
    # Price columns are at positions 8–13 (years 2019–2024), immediately after "Price on 1 April"
    # Revenue columns follow "Government revenue" at position 15 — we skip those here.
    price_april_idx = next(i for i, h in enumerate(gen_headers) if h == "Price on 1 April")
    price_year_indices = [(i, gen_headers[i]) for i in range(price_april_idx + 1, len(gen_headers)) if isinstance(gen_headers[i], int)]
    # Stop at "Change" (non-int) — only take the first block of ints
    first_block = []
    for i, yr in price_year_indices:
        if first_block and yr <= first_block[-1][1]:
            break
        first_block.append((i, yr))
    price_year_indices = first_block

    initiatives = []
    for row in gen_rows[2:]:
        if not row[0]:
            continue
        cpi_id = row[0]
        name = row[1]
        cpi_type = row[2]
        status = row[3]
        jurisdiction = row[4]
        share_jurisdiction = row[5]
        share_global = row[6]
        price_label = row[7]  # e.g. "€93.02 (US$99.99)*"

        initiatives.append(
            {
                "id": cpi_id,
                "name": name,
                "type": cpi_type,
                "status": status,
                "jurisdiction": jurisdiction,
                "share_jurisdiction_pct": round(share_jurisdiction * 100, 4) if isinstance(share_jurisdiction, float) else "",
                "share_global_pct": round(share_global * 100, 6) if isinstance(share_global, float) else "",
                "current_price_label": price_label or "",
            }
        )

    # --- Sheet 2: Compliance_Price → annual prices in USD ---
    ws_price = wb["Compliance_Price"]
    price_rows = list(ws_price.iter_rows(values_only=True))
    price_headers = price_rows[1]
    year_indices = [(i, h) for i, h in enumerate(price_headers) if isinstance(h, int)]

    price_records = []
    for row in price_rows[2:]:
        if not row[0]:
            continue
        name = row[0]
        cpi_type = row[1]
        jurisdiction = row[2]
        region = row[3]
        income_group = row[4]
        start_date = row[5]

        for col_idx, year in year_indices:
            val = row[col_idx] if col_idx < len(row) else None
            if val in (None, "", "-"):
                continue
            try:
                price = float(val)
            except (ValueError, TypeError):
                continue
            price_records.append(
                {
                    "initiative": name,
                    "type": cpi_type,
                    "jurisdiction": jurisdiction,
                    "region": region,
                    "income_group": income_group,
                    "start_year": start_date,
                    "year": year,
                    "price_usd_per_tco2e": round(price, 4),
                }
            )

    return initiatives, price_records


def fetch_nzu(url):
    raw = download_bytes(url).decode("utf-8")
    reader = csv.DictReader(io.StringIO(raw))
    records = []
    for row in reader:
        records.append({"date": row["date"].strip('"'), "price_nzd": row["price"]})
    return records


def write_csv(path, fieldnames, rows):
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)
    print(f"  Wrote {len(rows)} rows → {path}")


def main():
    print("Downloading World Bank Carbon Pricing Dashboard data...")
    wb_data = download_bytes(WB_URL)

    print("Parsing Excel...")
    initiatives, price_records = parse_wb_excel(wb_data)

    print("Downloading NZ ETS monthly prices...")
    nzu_records = fetch_nzu(NZU_URL)

    print("Writing CSVs...")

    # initiatives overview
    init_fields = list(initiatives[0].keys()) if initiatives else []
    write_csv("data/carbon-pricing-initiatives.csv", init_fields, initiatives)

    # annual prices (long format)
    price_fields = ["initiative", "type", "jurisdiction", "region", "income_group", "start_year", "year", "price_usd_per_tco2e"]
    write_csv("data/carbon-prices.csv", price_fields, price_records)

    # NZ ETS monthly
    write_csv("data/nz-ets-monthly-prices.csv", ["date", "price_nzd"], nzu_records)

    print("Done.")


if __name__ == "__main__":
    main()
